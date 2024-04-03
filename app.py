import os
from datetime import datetime
from time import sleep, time
import shutil

# import pyautogui
from openpyxl import load_workbook
from selenium import webdriver
from selenium.common.exceptions import *
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service as ChromeService
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support import expected_conditions as ec
from selenium.webdriver.support.wait import WebDriverWait
from webdriver_manager.chrome import ChromeDriverManager
from emailsender import Emailer
from config import EMAIL_ADDRESS, EMAIL_PASSWORD, EMAIL_ACESSO, SENHA_ACESSO, DIRETORIO_FILES


tempo_inicial = time()

# Diretório de Download de Faturas - Usado no Inicializar Driver
diretorio_raiz = os.path.join(os.getcwd(), str(datetime.today().year))
# Carregar a planilha com os dados das instalações
arquivo_faturas = "files/FATURAS.xlsx"
arquivo_resultado_downloads = f"files/Resultado{datetime.today().month}.xlsx"
wb = load_workbook(arquivo_faturas)
global ws
ws = wb.active
sleep(3)


def verifica_existencia_arquivo_resultado():
    global aba_resultado, resultado
    try:
        if os.path.isfile(arquivo_resultado_downloads):
            resultado = load_workbook(arquivo_resultado_downloads)
            aba_resultado = resultado.active
            return aba_resultado
        else:
            shutil.copyfile(arquivo_faturas, arquivo_resultado_downloads)
            resultado = load_workbook(f"files/Resultado{datetime.today().month}.xlsx")
            aba_resultado = resultado.active
            return aba_resultado
    except FileNotFoundError:
        print("-----")

    tempo_inicial = time()


def enviar_mail(tempo, erro=None):
    email = Emailer(EMAIL_ADDRESS, EMAIL_PASSWORD)
    lista_contatos = ["a.alves@perkons.com"]
    if erro:
        mensagem = f"""
        O script de automação para EDP foi finalizado falha.
        Tempo de Execução: {tempo}

        exceção: {erro.__str__()}
        """
        topico = "Deu Falha no Script Automação EDP"
    else:
        mensagem = f"""
        O script de automação para EDP foi finalizado com sucesso.
        Tempo de Execução: {tempo}"""

        topico = "Funcionou corretamente"

    email.definir_conteudo(
        topico=topico,
        email_remetente="andre@andrealves.eng.br",
        lista_contatos=lista_contatos,
        conteudo_email=mensagem,
    )
    email.anexar_arquivos([f"files/Resultado{datetime.today().month}.xlsx"])

    email.enviar_email(intervalo_em_segundos=5)


def iniciar_driver():
    chrome_options = Options()
    arguments = ["--lang=pt-BR", "--start-maximized", "--enable-javascript"]
    for argument in arguments:
        chrome_options.add_argument(argument)

    chrome_options.add_experimental_option('prefs', {
        'download.default_directory': os.path.join(diretorio_raiz, str(datetime.today().month)),
        'download.directory_upgrade': True,
        'download.prompt_for_download': False,
        'profile.default_content_setting_values.notifications': 2,
        'profile.default_content_setting_values.automatic_downloads': 1,

    })
    navegador = webdriver.Chrome(service=ChromeService(
        ChromeDriverManager().install()), options=chrome_options)

    espera = WebDriverWait(
        navegador,
        60,
        poll_frequency=1,
        ignored_exceptions=[
            NoSuchElementException,
            ElementNotVisibleException,
            ElementNotSelectableException,
        ],
    )
    return navegador, espera


def acessar_o_site():
    global driver
    driver, wait = iniciar_driver()
    driver.get("https://www.edponline.com.br/servicos")

    # aceitar cookies
    sleep(2)
    driver.find_element(
        By.XPATH,
        "//div[@id='onetrust-button-group']/button[@id='onetrust-accept-btn-handler']",
    ).click()

    # clica na aba meus negócios.
    sleep(2)
    driver.find_element(
        By.XPATH, '//div[@class="login-layout__login-form__tab-nav"]/a[2]'
    ).click()

    # Escolher Espírito Santo.
    (
        WebDriverWait(driver, 60)
        .until(
            ec.element_to_be_clickable(
                (By.XPATH, '//*[@id="login-form"]/div[1]/div[2]')
            )
        )
        .click()
    )
    sleep(1)

    # Entrar com usuário e senha
    (
        WebDriverWait(driver, 60)
        .until(ec.presence_of_element_located((By.XPATH, '//input[@name="Email"]')))
        .send_keys(EMAIL_ACESSO)
    )
    sleep(1)

    (
        WebDriverWait(driver, 60)
        .until(ec.presence_of_element_located((By.XPATH, '//input[@name="Senha"]')))
        .send_keys(SENHA_ACESSO)
    )
    sleep(1)

    # Clicar para acessar o site.
    WebDriverWait(driver, 60).until(
        ec.presence_of_element_located((By.XPATH, '//button[@id="acessar"]'))
    ).click()


def verifica_download_realizado(index):
    # Lê na planilha qual planilha já foi feito download e pula essa linha.
    if aba_resultado.cell(row=index, column=4).value == "Download Realizado":
        print("Fatura já baixada !")
        return True


def tratar_excecao(excecao, index):
    # Escreve na planilha que o download caso de exeção
    aba_resultado.cell(row=index, column=4, value="Download Não realizado")
    resultado.save(f"files/Resultado{datetime.today().month}.xlsx")
    tempo_final = time()
    tempo = tempo_final - tempo_inicial
    enviar_mail(tempo, excecao)


def realizar_download():
    for index, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
        print(row)
        # Função para verificar se a fatura já foi baixada
        if verifica_download_realizado(index):
            continue

        # Limpar o campo e pesquisar número da instalação
        (
            WebDriverWait(driver, 60)
            .until(ec.presence_of_element_located((By.XPATH, '//input[@id="dados"]')))
            .clear()
        )
        sleep(1)
        (
            WebDriverWait(driver, 60)
            .until(ec.presence_of_element_located((By.XPATH, '//input[@id="dados"]')))
            .send_keys(row[1])
        )
        sleep(1)
        (
            WebDriverWait(driver, 60)
            .until(ec.presence_of_element_located((By.XPATH, '//input[@id="dados"]')))
            .send_keys(Keys.RETURN)
        )
        sleep(1)
        # checar qual status da instalação
        status = WebDriverWait(driver, 60).until(
            ec.presence_of_element_located(
                (By.XPATH, '//*[@id="grid"]/table/tbody/tr/td[2]')
            )
        )
        if status.text == "CONTRATO ENCERRADO":
            continue
        # clicar no link do número de instalação
        (
            WebDriverWait(driver, 60)
            .until(
                ec.presence_of_element_located((By.XPATH, "//table/tbody/tr/td[1]/a"))
            )
            .click()
        )
        sleep(2)
        # Clicar visualizar ultimas contas
        (
            WebDriverWait(driver, 60)
            .until(
                ec.presence_of_element_located(
                    (By.XPATH, '//*[@id="ultima-conta"]/div/div[2]/div[3]/a[2]')
                )
            )
            .click()
        )
        sleep(1)
        try:
            # Clicar em ver faturas
            (
                WebDriverWait(driver, 60)
                .until(
                    ec.presence_of_element_located(
                        (
                            By.XPATH,
                            '//*[@id="extrato-de-contas"]/div/div/div/div[1]'
                            "/div/div/div/div[1]/div/div/div[3]/div[2]/div[2]/a[1]",
                        )
                    )
                )
                .click()
            )
            # Clicar para baixar pdf.
            (
                WebDriverWait(driver, 60)
                .until(
                    ec.presence_of_element_located(
                        (
                            By.XPATH,
                            '//div[@id="box-dados-fatura"]//a[@class="i-block text-center pull-right"]',
                        )
                    )
                )
                .click()
            )
            sleep(3)
            # Fecha a janela da fatura
            (
                WebDriverWait(driver, 60)
                .until(
                    ec.presence_of_element_located(
                        (By.XPATH, '//a[@data-dismiss="modal"]')
                    )
                )
                .click()
            )
            sleep(3)
            # Escreve na planilha que o download já foi realizado
            aba_resultado.cell(row=index, column=4, value="Download Realizado")
            driver.get("https://www.edponline.com.br/servicos/selecionar-instalacao")
            resultado.save(f"files/Resultado{datetime.today().month}.xlsx")
            sleep(5)
        except ElementNotInteractableException:
            tratar_excecao(excecao=ElementNotInteractableException, index=index)
            sleep(5)
            driver.get("https://www.edponline.com.br/servicos/selecionar-instalacao")
            continue
        except StaleElementReferenceException:
            tratar_excecao(excecao=StaleElementReferenceException, index=index)
            sleep(5)
            driver.get("https://www.edponline.com.br/servicos/selecionar-instalacao")
            continue
        except TimeoutException:
            tratar_excecao(excecao=TimeoutException, index=index)
            sleep(5)
            driver.get("https://www.edponline.com.br/servicos/selecionar-instalacao")
            continue


verifica_existencia_arquivo_resultado()
iniciar_driver()
acessar_o_site()
realizar_download()


tempo_final = time()
tempo = tempo_final - tempo_inicial
enviar_mail(tempo)
